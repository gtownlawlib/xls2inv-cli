# Import modules/packages
import sys
import time
from openpyxl import load_workbook

# Load worksheet
wb2 = load_workbook(sys.argv[1])
ws1 = wb2['Sheet1']

# Begin invoice header string
output = 'H'

# Set invoice number based on staff user
headInv = ws1.cell(row=2, column=8)
output += '%s' % headInv.value

# Set invoice header date
nnm = time.strftime('%m')
nnd = time.strftime('%d')
nny = time.strftime('%y')

output += '%s%s%s' % (nnm, nnd, nny)

# Add up subtotal of all line items (pre-tax/pre-S&H) and add to header as
# subtotal string (nine characters, right justified, zero fill, decimal
# point implied)
headSubtotal = 0
rowcount = 2
for row in ws1.rows:
    price = ws1.cell(row=rowcount, column=4)
    rowcount = rowcount + 1
    if price.value is not None:
        headSubtotal = headSubtotal + price.value
headSubtotal = str(headSubtotal).replace('.', '')
headSubtotal = headSubtotal.zfill(9)

output += '%s' % headSubtotal

# Add string of 9 zeroes for header service charge because service charges
# are merged with tax in spreadsheet

output += '0' * 9

# Add up S&H/tax charges of all line items and add to header as tax string
# (nine characters, right justified, zero fill, decimal point implied)
headTax = 0
rowcount = 2
for row in ws1.rows:
    tax = ws1.cell(row=rowcount, column=5)
    rowcount = rowcount + 1
    if tax.value is not None:
        headTax = headTax + tax.value

headTax = str(headTax).replace('.', '')
headTax = headTax.zfill(9)

output += '%s' % headTax

# Add subtotal and tax and add to header as grandtotal string (nine
# characters, right justified, zero fill, decimal point implied)
headSubtotal = 0
rowcount = 2
for row in ws1.rows:
    price = ws1.cell(row=rowcount, column=4)
    rowcount = rowcount + 1
    if price.value is not None:
        headSubtotal = headSubtotal + price.value

headTax = 0
rowcount = 2
for row in ws1.rows:
    tax = ws1.cell(row=rowcount, column=5)
    rowcount = rowcount + 1
    if tax.value is not None:
        headTax = headTax + tax.value

headGrand = headSubtotal + headTax
headGrand = str(headGrand).replace('.', '')
headGrand = headGrand.zfill(9)

output += '%s' % headGrand

# Add number of rows to header as number of line items string (five digits,
# right justified, zero fill, max value of 500)

for row in ws1.rows:
    actualVals = [cell.value for cell in row]
    if any(actualVals):
        headItems = '{0}'.format(row[0].row)
headItems = int(headItems) - 1
headItems = str(headItems).zfill(5)

output += '%s' % headItems

# End header with blank space and asterisk then insert line break
output += ' ' * 44
output += '*'
output += '\n'

# Begin iterating through worksheet rows to generate line items
rowcount = 2
for row in ws1.rows:

    # Set column variables
    invdate = ws1.cell(row=rowcount, column=1)  # 'Order Date' column
    ordnum = ws1.cell(row=rowcount, column=2)  # 'Order Number' column
    numcopy = ws1.cell(row=rowcount, column=3)  # '# of Copy' column
    amount = ws1.cell(row=rowcount, column=6)  # 'Total Cost' column
    note = ws1.cell(row=rowcount, column=7)  # 'Note' column

    # Don't process a row unless an order number is present
    if ordnum.value is not None:
        # Insert L at beginning of each line item
        output += 'L'

        # Fill in (dummy?) invoice number and date
        if invdate.value is not None:

            # Convert date column to time object
            invdatestr = time.strptime(
                str(invdate.value),
                '%Y-%m-%d %H:%M:%S'
            )

            # Format invoice number and date from object
            invnum = str(time.strftime('%j%Y', invdatestr))
            invdateval = str(time.strftime('%m%d%y', invdatestr))
            output += '%s%s' % (invnum, invdateval)

        # Fill in Innovative record (i.e., order) #; strip leading 'o'
        if ordnum.value is not None:
            ordnumval = str(ordnum.value)
            if ordnumval.startswith('o') is True:
                ordnumval = ordnumval[1:]
            ordnumval = ordnumval.ljust(8)
            output += ordnumval

        # Blank fill
        output += ' ' * 11

        # Map note column to vendor subscription ID
        output += '!'
        if note.value is not None:
            note.value = note.value[:29]
            output += note.value.ljust(29)
        else:
            output += ' ' * 29  # Dummy value

        # Fill in amount
        if amount.value is not None:
            amtval = str(amount.value)
            amtval = amtval.replace('.', '')
            amtval = amtval.zfill(7)
            output += amtval

        # Fill in number of copies (# of copies column)
        if numcopy.value is not None:
            numval = str(numcopy.value)
            numval = numval.zfill(3)
            output += numval

        # Fill in line-level tax (GST/VAT)
        output += '0' * 7

        # Blank fill
        output += ' ' * 7

        # Fill subscription dates (order date to order date plus 1 yr)
        if invdate.value is not None:
            substartint = int(invdateval)
            subend = str(substartint + 1).zfill(6)
            substart = str(substartint).zfill(6)
            output += '%s%s' % (substart, subend)

        # End line item with an asterisk and line break
        output += '*'
        output += '\n'

    # Iterate the rowcount
    rowcount = rowcount + 1

# Write data to output file
f = open(sys.argv[2], 'w+')
f.write('%s' % output)
f.close()
